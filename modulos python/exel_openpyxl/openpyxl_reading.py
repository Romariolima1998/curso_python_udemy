from pathlib import Path
from openpyxl import load_workbook

root_folder = Path(__file__).parent
workbook_path = root_folder / 'workbook.xlsx'

# carregando um arquivo excel
workbook = load_workbook(workbook_path)

sheet_name = 'minha planilha'

worksheet = workbook[sheet_name]

for row in worksheet.iter_rows():
    print()
    for cell in row:
        print(cell.value)
        if cell.value == 'maria':
            worksheet.cell(cell.row, 2, '15')
            
# alterar o valor
# worksheet['B3'].value = 16

workbook.save(workbook_path)