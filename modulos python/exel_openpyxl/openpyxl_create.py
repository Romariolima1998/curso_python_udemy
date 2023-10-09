# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

root_folder = Path(__file__).parent
workbook_path = root_folder / 'workbook.xlsx'

# nome da planilha
sheet_name = 'minha planilha'
workbook = Workbook()
# worksheet: Worksheet = workbook.active
# criando a planilha
workbook.create_sheet(sheet_name, 0)
# selecionar planilha
worksheet: Worksheet = workbook[sheet_name]
# remover uma planilha
workbook.remove(workbook['Sheet'])
print(workbook.sheetnames)

studets = [
    # nome idade nota
    ['joao', 14, 5.5],
    ['maria', 13, 9.7],
    ['luiz', 15, 8.8],
    ['alberto', 16, 10],
]

# criando cabeçalho
worksheet.cell(1, 1, 'nome')
worksheet.cell(1, 2, 'idade')
worksheet.cell(1, 3, 'nota')

# for i, student_row in enumerate(studets, start=2):
#    for j, student_column in enumerate(student_row, start=1):
#        worksheet.cell(i, j, student_column)

for student in studets:
    worksheet.append(student)
workbook.save(workbook_path)
